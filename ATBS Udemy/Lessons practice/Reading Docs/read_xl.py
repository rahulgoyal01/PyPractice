from openpyxl import Workbook, load_workbook
import os

#file should be present in CWD
#wb = Workbook() 
wb = load_workbook(filename = 'Process Log.xlsx')
print(wb.sheetnames)
sheet = wb['plogs']
print(sheet['A2'].value)
print(sheet['b2'].value)
print(sheet.cell(row=2, column=1))
print(sheet.cell(row=2, column=1).value)

print('Below are the first 8 processes')
for i in range(2,10):
    
    print((i-1), sheet.cell(row=i, column=1).value)
